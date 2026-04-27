import json
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

excel_path = r"C:\Users\travi\Origin-PXRD\build\option_files\Annealing Profiles\Furnaces\Furnace.xlsx"

wb = load_workbook(excel_path, data_only=True)

furnaces = {}
two_point_bools = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # ws.tables is a dict: {table_name: TableObject}
    for table_name, table_obj in ws.tables.items():
        ref = table_obj # e.g. "A1:B10"

        

        

        # Convert Excel range to row/col indices
        min_col, min_row, max_col, max_row = range_boundaries(ref)

        # Read the entire sheet once
        df_sheet = pd.read_excel(excel_path, sheet_name=sheet_name, header=0)

        # Slice the table range (convert to 0‑based indexing)
        df_table = df_sheet.iloc[min_row-1:max_row, min_col-1:max_col]

        # Clean up
        df_table = df_table.dropna(how="all")

        if table_name == "interpolate":
            id_col = df_table.columns[0]
            two_point_col = df_table.columns[3]
            for furnace_id, two_point in zip(df_table[id_col], df_table[two_point_col]):
                two_point_bools[f'{furnace_id:02d}'] = two_point=='Y'

        if "FURNACE" not in table_name.upper():
            continue

        id_num = re.search(r"(\d+)$", table_name).group(1)
        key = "FRN" + id_num

        # Extract the two columns
        set_col = df_table.columns[0]
        actual_col = df_table.columns[1]

        furnaces[key] = {
            "set": df_table[set_col].dropna().tolist(),
            "actual": df_table[actual_col].dropna().tolist(),
            "two_point": two_point_bools.get(id_num, False)
        }

# Save JSON
with open("furnace_calibrations.json", "w") as f:
    json.dump(furnaces, f, indent=4)

print("Done.")
